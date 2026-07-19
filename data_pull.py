# data_pull.py
# Pulls ETF price data via yfinance and the 3-Month T-Bill rate via fredapi.
# Constructs the fixed multi-tab Excel workbook used as input across all trial
# conditions. Run once before any trials; do not re-run between conditions.
#
# Output: data/etf_returns.xlsx
#   - One tab per ETF ticker containing: Date, Price, Return columns
#   - One "RF" tab containing: Date, RF_Annual_Pct, RF_Periodic_Decimal columns
#
# Usage:
#   python data_pull.py               # full run (all tickers)
#   python data_pull.py --tabs 3      # test run (first 3 tickers only)

import argparse
import os

import pandas as pd
import yfinance as yf
from fredapi import Fred
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import (
    TICKERS,
    START_DATE,
    END_DATE,
    RETURN_FREQUENCY,
    ANNUALIZATION_FACTOR,
    FRED_SERIES_ID,
    FRED_API_KEY,
    WORKBOOK_PATH,
)
from utils import convert_rf_to_periodic


# =============================================================================
# ARGUMENT PARSING
# =============================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Pull ETF and risk-free rate data and build input workbook."
    )
    parser.add_argument(
        "--tabs",
        type=int,
        default=None,
        metavar="N",
        help="Limit to first N tickers for test runs. Omit for full run."
    )
    return parser.parse_args()


# =============================================================================
# ETF DATA PULL
# =============================================================================

def pull_etf_prices(tickers: list[str]) -> dict[str, pd.Series]:
    """
    Download adjusted closing prices for each ticker via yfinance.
    Resamples to month-end if RETURN_FREQUENCY == "monthly".

    Parameters
    ----------
    tickers : list[str]
        List of ETF ticker symbols.

    Returns
    -------
    dict[str, pd.Series]
        Mapping of ticker -> price series with DatetimeIndex.
    """
    print(f"Pulling price data for {len(tickers)} tickers ({START_DATE} to {END_DATE})...")

    raw = yf.download(
        tickers=tickers,
        start=START_DATE,
        end=END_DATE,
        auto_adjust=True,
        progress=False,
    )

    # yfinance returns MultiIndex columns when multiple tickers are passed
    if isinstance(raw.columns, pd.MultiIndex):
        prices_df = raw["Close"]
    else:
        prices_df = raw[["Close"]]
        prices_df.columns = tickers

    if RETURN_FREQUENCY == "monthly":
        prices_df = prices_df.resample("ME").last()

    price_dict = {}
    for ticker in tickers:
        series = prices_df[ticker].dropna()
        if len(series) < 2:
            print(f"  WARNING: Insufficient data for {ticker} — skipping.")
            continue
        price_dict[ticker] = series
        print(f"  {ticker}: {len(series)} observations")

    return price_dict


# =============================================================================
# RISK-FREE RATE PULL
# =============================================================================

def pull_risk_free_rate() -> tuple[pd.Series, pd.Series]:
    """
    Pull the 3-Month T-Bill rate from FRED and convert to per-period decimal.

    Returns
    -------
    tuple[pd.Series, pd.Series]
        (rf_annual_pct, rf_periodic_decimal)
        rf_annual_pct    : raw FRED series (annualized percentages)
        rf_periodic_decimal : per-period decimal rates aligned to return frequency
    """
    print(f"Pulling risk-free rate from FRED (series: {FRED_SERIES_ID})...")

    fred = Fred(api_key=FRED_API_KEY)
    rf_raw = fred.get_series(
        FRED_SERIES_ID,
        observation_start=START_DATE,
        observation_end=END_DATE,
    )
    rf_raw.name = "RF_Annual_Pct"
    rf_raw.index = pd.to_datetime(rf_raw.index)

    if RETURN_FREQUENCY == "monthly":
        rf_raw = rf_raw.resample("ME").last().ffill()

    rf_periodic = convert_rf_to_periodic(rf_raw)
    rf_periodic.name = "RF_Periodic_Decimal"

    print(f"  Risk-free rate: {len(rf_raw)} observations")
    return rf_raw, rf_periodic


# =============================================================================
# WORKBOOK CONSTRUCTION
# =============================================================================

def _write_header_row(ws, headers: list[str]) -> None:
    """Apply consistent header formatting across all tabs."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


def write_etf_tab(ws, ticker: str, prices: pd.Series) -> None:
    """
    Write price and return data for one ETF to a worksheet tab.

    Columns: Date | Price | Return
    Return is simple periodic return expressed as a decimal.
    First row has NaN return (no prior price); written as blank.

    Parameters
    ----------
    ws : openpyxl Worksheet
        Target worksheet.
    ticker : str
        ETF ticker symbol (used in tab header comment).
    prices : pd.Series
        Month-end adjusted closing prices.
    """
    returns = prices.pct_change()
    _write_header_row(ws, ["Date", "Price", "Return"])

    for row_idx, (date, price) in enumerate(prices.items(), start=2):
        ret = returns.loc[date]
        ws.cell(row=row_idx, column=1, value=date.strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=2, value=round(float(price), 6))
        ws.cell(row=row_idx, column=3, value=round(float(ret), 6) if pd.notna(ret) else None)


def write_rf_tab(ws, rf_annual: pd.Series, rf_periodic: pd.Series) -> None:
    """
    Write risk-free rate data to the RF tab.

    Columns: Date | RF_Annual_Pct | RF_Periodic_Decimal
    Provides full transparency on rate conversion for methodology verification.

    Parameters
    ----------
    ws : openpyxl Worksheet
        Target worksheet.
    rf_annual : pd.Series
        Annualized percentage rates from FRED.
    rf_periodic : pd.Series
        Per-period decimal rates after conversion.
    """
    _write_header_row(ws, ["Date", "RF_Annual_Pct", "RF_Periodic_Decimal"])

    for row_idx, (date, annual_rate) in enumerate(rf_annual.items(), start=2):
        periodic_rate = rf_periodic.loc[date] if date in rf_periodic.index else None
        ws.cell(row=row_idx, column=1, value=date.strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=2, value=round(float(annual_rate), 6) if pd.notna(annual_rate) else None)
        ws.cell(row=row_idx, column=3, value=round(float(periodic_rate), 8) if periodic_rate is not None and pd.notna(periodic_rate) else None)


def build_workbook(
    price_dict: dict[str, pd.Series],
    rf_annual: pd.Series,
    rf_periodic: pd.Series,
    output_path: str,
) -> None:
    """
    Assemble the multi-tab Excel workbook and write to disk.

    Tab order: one tab per ETF (in TICKERS order) followed by RF tab.

    Parameters
    ----------
    price_dict : dict[str, pd.Series]
        Ticker -> price series mapping.
    rf_annual : pd.Series
        Annualized percentage risk-free rates.
    rf_periodic : pd.Series
        Per-period decimal risk-free rates.
    output_path : str
        Destination file path for the workbook.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    for ticker, prices in price_dict.items():
        ws = wb.create_sheet(title=ticker)
        write_etf_tab(ws, ticker, prices)
        print(f"  Written tab: {ticker}")

    ws_rf = wb.create_sheet(title="RF")
    write_rf_tab(ws_rf, rf_annual, rf_periodic)
    print(f"  Written tab: RF")

    wb.save(output_path)
    print(f"\nWorkbook saved to: {output_path}")
    print(f"Tabs: {[ws.title for ws in wb.worksheets]}")


# =============================================================================
# MAIN
# =============================================================================

def main() -> None:
    args = parse_args()

    tickers = TICKERS[:args.tabs] if args.tabs is not None else TICKERS

    if args.tabs is not None:
        print(f"TEST RUN: limiting to first {args.tabs} tickers.\n")
    else:
        print("FULL RUN: processing all tickers.\n")

    price_dict = pull_etf_prices(tickers)
    rf_annual, rf_periodic = pull_risk_free_rate()
    build_workbook(price_dict, rf_annual, rf_periodic, WORKBOOK_PATH)

    print("\ndata_pull.py complete.")
    print(f"  ETF tabs written : {len(price_dict)}")
    print(f"  RF tab written   : 1")
    print(f"  Output path      : {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
